import feedparser
import csv
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime
import os

# Define a function to print country codes
def print_country_codes():
    country_codes = """
    RASFF Country Codes (Europe)
    
    AUSTRIA 5001
    BELGIUM 5002
    BULGARIA 5003
    CROATIA 5016
    CYPRUS 5005
    CZECH REPUBLIC 5006
    DENMARK 5008
    ESTONIA 5009
    FINLAND 5012
    FRANCE 5013
    GERMANY 5007
    GREECE 5015
    HUNGARY 5017
    ICELAND 5019
    IRELAND 5018
    ITALY 5020
    LATVIA 5024
    LIECHTENSTEIN 5021
    LITHUANIA 5022
    LUXEMBOURG 5023
    MALTA 5025
    NETHERLANDS 5026
    NORWAY 5027
    POLAND 5028
    PORTUGAL 5029
    ROMANIA 5030
    SLOVAKIA 5033
    SLOVENIA 5032
    SPAIN 5010
    SWEDEN 5031
    SWITZERLAND 5004
    
    Usage: RASFF --country CODE (or no code for all EU Single Market notifications)
    """
    print(country_codes)
    
# Function to parse command-line arguments
def parse_arguments():
    parser = argparse.ArgumentParser(description='Fetch notifications from RSS feed and save to CSV.')
    parser.add_argument('--country', type=str, default='all', help='Country code for the RSS feed (default: "all").')
    parser.add_argument('--list-countries', action='store_true', help='List the country codes and exit.')

    args = parser.parse_args()

    if args.list_countries:
        print_country_codes()
        parser.exit()  # Exit after printing country codes
    
    return args

args = parse_arguments()

# Function to fetch and parse the RSS feed
def fetch_rss_feed(url):
    return feedparser.parse(url)

# Function to extract information from each entry in the feed
def process_feed_entries(entries):
    processed_data = []
    for entry in entries:
        data = {
            'title': entry.title,
            'link': entry.link,
            'notified': entry.description,
            # Extract more fields as needed
        }
        processed_data.append(data)
    return processed_data

# Function to write the data to a CSV file and print lines on-screen
def write_to_csv(data, filename='notifications.csv'):
    with open(filename, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(['Title', 'Link', 'Notified'])  # Adjust the header based on the fields you're extracting
        for item in data:
            writer.writerow([item['title'], item['link'], item['notified']])  # Adjust accordingly
            print(f'Title: {item["title"]}, Link: {item["link"]}, Notified: {item["notified"]}')  # Print on-screen
            
from openpyxl import Workbook

def write_to_xlsx(data, filename='notifications.xlsx'):
    # Check if the file exists
    if os.path.exists(filename):
        try:
            # Try to load the existing workbook
            wb = load_workbook(filename)
        except InvalidFileException:
            # If the file is not a valid .xlsx file, create a new workbook
            wb = Workbook()
    else:
        # If the file does not exist, create a new workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet

    # Generate sheet name based on current date
    sheet_name = f"Notifications_{datetime.now().strftime('%Y%m%d')}"

    # Create a new sheet with today's date
    ws = wb.create_sheet(title=sheet_name)

    # Headers
    ws.append(['Title', 'Link', 'Notified'])

    # Data
    for item in data:
        ws.append([item['title'], item['link'], item['notified']])
        print(f'Title: {item["title"]}, Link: {item["link"]}, Notified: {item["notified"]}')  # Print on-screen

    # Save the workbook
    wb.save(filename)

# Main function to tie it all together
def main():
    args = parse_arguments()
    country_code = args.country
    rss_url = f"https://webgate.ec.europa.eu/rasff-window/backend/public/consumer/rss/{country_code}/"
    feed = fetch_rss_feed(rss_url)
    processed_data = process_feed_entries(feed.entries)
    #output_filename = f'notifications-{country_code}.doc'
    #write_to_csv(processed_data, output_filename)
    
    output_filename = f'notifications-{country_code}.xlsx'
    write_to_xlsx(processed_data, output_filename)
    print(f"RSS feed data has been written to {output_filename}")
    
    
    # Attribution message
    print("Consulta RASFF, Carlos Simoes & OpenAI, Abril 2024")


if __name__ == "__main__":
    main()



